# Esto ayuda a crear nuevos espacios desde cero
from openpyxl import Workbook, load_workbook
import openpyxl
from openpyxl.drawing.image import Image
#Esto ayuda a convertir columnas en numeros
from openpyxl.utils import get_column_letter, column_index_from_string
################################
# wb = openpyxl.load
# ws = wb.active
# ws1 = wb.create_sheet('MySheet', 0)
# ws2 = wb.create_sheet('MySheet')
# for sheet in wb:
#     sheet.title="MySheet"
# # Manipulando Datos
# a = wb['MySheet']['A1':'A10']
# wb.save("NombreArchivo.xlsx")
#################################

###################################
# Manipulando Archivos excel
# wb = openpyxl.Workbook() # otra forma sin importar
file_name ='test.xlsx'

wb = load_workbook('sample.xlsx')

ws = wb.active

print(wb.sheetnames)

sheet = wb['Sheet3']

print(sheet.title)

###############################################
# Iteracion para cada celda
#
# anotherSheet = wb.active
# for i in anotherSheet['A1':'C8']:
#     print("[",end=" ")
#     for j in i:
#         print(j.value, end=" ")
#     print("],")
###############################################


sheet1 = wb['Sheet']
# Otra Forma de iterar por la matriz mas sencillo
# for i in range(1,5):
#     print(i, sheet1.cell(row=i, column=2).value)

# print(sheet1.max_row)
# print(sheet1.max_column)

# # Despliega Letra de Columna
# print(get_column_letter(1))
# print(get_column_letter(900))

# print("Max column:", get_column_letter(sheet.max_column))

# print(column_index_from_string('AA'))


###############################################
# Iteracion de otra forma:
# itera = wb['Sheet']['A1':'C3']
# cont =1
# for rowOfCellObjects in sheet['A1':'C3']:
#     for cellObj in rowOfCellObjects:
#         cellObj.value = cont
#         print(cellObj.coordinate, cellObj.value)
#         cont+=1
#     print('----------')

############################################
# # Crear sheets nuevos entre otros
# wb.create_sheet(index=2, title='Middle Sheet') 
# print(wb.sheetnames)

################################################
# # Borrar Sheets
# del wb['Middle Sheet']
# print(wb.sheetnames)

################################################
# # Para hacer un merge se necesita activar en una
# # variable y luego hacer uso. Combinar celdas
# ws.merge_cells('A5:A6')

################################################
# # Agregar una imagen. Primero se debe importar
# # la libreria 
# # from openpyxl.drawing.image import Image

# img = Image('graphTemp.png')
# ws.add_image(img, 'A8')

##################################################
# # Fold (outline)
# ws.column_dimensions.group('G','M', hidden=True)
# ws.row_dimensions.group(30,35, hidden=True)


#################################################
# # Deleting rows and columns
# # To delete the columns F:H:
# ws.delete_cols(6, 3)

#################################################
# # Moving ranges of cells
# # You can also move ranges of cells within a worksheet:
# ws.move_range("D4:F10", rows=-1, cols=2)

#################################################
# parsing formulas transalator:
# ws.move_range("G4:H10", rows=1, cols=1, translate=True)



#################################################
# working with pandas dataframes
# from openpyxl.utils.dataframe import dataframe_to_rows
# for r in dataframe_to_rows(df, index=True, header=True):
#     ws.append(r)
# for cell in ws['A'] + ws[1]:
#     cell.style = 'Pandas'


#################################################
# Alternatively, if you just want to convert the data you can use write-only mode:
# from openpyxl.cell.cell import WriteOnlyCell
# wb = Workbook(write_only=True)
# ws = wb.create_sheet()

# cell = WriteOnlyCell(ws)
# cell.style = 'Pandas'

#  def format_first_row(row, cell):

#     for c in row:
#         cell.value = c
#         yield cell

# rows = dataframe_to_rows(df)
# first_row = format_first_row(next(rows), cell)
# ws.append(first_row)

# for row in rows:
#     row = list(row)
#     cell.value = row[0]
#     row[0] = cell
#     ws.append(row)


################################################
# Converting a worksheet to a Dataframe
# df = DataFrame(ws.values)

# # If the worksheet does have headers or indices, 
# # such as one created by Pandas, then a little 
# # more work is required:
#
# from itertools import islice
# data = ws.values
# cols = next(data)[1:]
# data = list(data)
# idx = [r[0] for r in data]
# data = (islice(r, 1, None) for r in data)
# df = DataFrame(data, index=idx, columns=cols)



################Creando un Grafico##############
# Creando un Grafico:
for i in range(10):
    ws.append([i])

from openpyxl.chart import BarChart, Reference, Series
values = Reference(ws, min_col=1, min_row=1, max_col=1, max_row=10)
chart = BarChart()
chart.add_data(values)
ws.add_chart(chart, "E15")

# By default the top-left corner of a chart is 
# anchored to cell E15 and the size is 15 x 7.5 cm 
# (approximately 5 columns by 14 rows). 
# This can be changed by setting the anchor, 
# width and height properties of the chart. 
# The actual size will depend on operating system 
# and device. Other anchors are possible; 
# see openpyxl.drawing.spreadsheet_drawing for 
# further information.



###################Graficos de Area###########################
# from openpyxl.chart import (
#     AreaChart,
#     Reference,
#     Series,
# )

# wb = Workbook()
# ws = wb.active

# rows = [
#     ['Number', 'Batch 1', 'Batch 2'],
#     [2, 40, 30],
#     [3, 40, 25],
#     [4, 50, 30],
#     [5, 30, 10],
#     [6, 25, 5],
#     [7, 50, 10],
# ]

# for row in rows:
#     ws.append(row)

# chart = AreaChart()
# chart.title = "Area Chart"
# chart.style = 13
# chart.x_axis.title = 'Tiempo'
# chart.y_axis.title = 'Percentage'

# cats = Reference(ws, min_col=1, min_row=1, max_row=7)
# data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=7)
# chart.add_data(data, titles_from_data=True)
# chart.set_categories(cats)

# ws.add_chart(chart, "A10")


###############Graficos de Area 3D ############################
# from openpyxl.chart import (
#     AreaChart3D,
#     Reference,
#     Series,
# )

# wb = Workbook()
# ws = wb.active

# rows = [
#     ['Number', 'Batch 1', 'Batch 2'],
#     [2, 30, 40],
#     [3, 25, 40],
#     [4 ,30, 50],
#     [5 ,10, 30],
#     [6,  5, 25],
#     [7 ,10, 50],
# ]

# for row in rows:
#     ws.append(row)

# chart = AreaChart3D()
# chart.title = "Area Chart"
# chart.style = 13
# chart.x_axis.title = 'Test'
# chart.y_axis.title = 'Percentage'
# chart.legend = None

# cats = Reference(ws, min_col=1, min_row=1, max_row=7)
# data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=7)
# chart.add_data(data, titles_from_data=True)
# chart.set_categories(cats)

# ws.add_chart(chart, "A10")


###################### Bar Charts ######################
# from openpyxl.chart import BarChart, Series, Reference

# wb = Workbook(write_only=True)
# ws = wb.create_sheet()

# rows = [
#     ('Number', 'Batch 1', 'Batch 2'),
#     (2, 10, 30),
#     (3, 40, 60),
#     (4, 50, 70),
#     (5, 20, 10),
#     (6, 10, 40),
#     (7, 50, 30),
# ]

# for row in rows:
#     ws.append(row)

# chart1 = BarChart()
# chart1.type = "col"
# chart1.style = 10
# chart1.title = "Bar Chart"
# chart1.y_axis.title = 'Test number'
# chart1.x_axis.title = 'Sample length (mm)'

# data = Reference(ws, min_col=2, min_row=1, max_row=7, max_col=3)
# cats = Reference(ws, min_col=1, min_row=2, max_row=7)
# chart1.add_data(data, titles_from_data=True)
# chart1.set_categories(cats)
# chart1.shape = 4
# ws.add_chart(chart1, "A10")

# from copy import deepcopy

# chart2 = deepcopy(chart1)
# chart2.style = 11
# chart2.type = "bar"
# chart2.title = "Horizontal Bar Chart"

# ws.add_chart(chart2, "G10")

# chart3 = deepcopy(chart1)
# chart3.type = "col"
# chart3.style = 12
# chart3.grouping = "stacked"
# chart3.overlap = 100
# chart3.title = 'Stacked Chart'

# ws.add_chart(chart3, "A27")

# chart4 = deepcopy(chart1)
# chart4.type = "bar"
# chart4.style = 13
# chart4.grouping = "percentStacked"
# chart4.overlap = 100
# chart4.title = 'Percent Stacked Chart'

# ws.add_chart(chart4, "G27")

######################## 3D Bar Charts #######################
# from openpyxl.chart import (
#     Reference,
#     Series,
#     BarChart3D,
# )

# wb = Workbook()
# ws = wb.active

# rows = [
#     (None, 2013, 2014),
#     ("Apples", 5, 4),
#     ("Oranges", 6, 2),
#     ("Pears", 8, 3)
# ]

# for row in rows:
#     ws.append(row)

# data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=4)
# titles = Reference(ws, min_col=1, min_row=2, max_row=4)
# chart = BarChart3D()
# chart.title = "3D Bar Chart"
# chart.add_data(data=data, titles_from_data=True)
# chart.set_categories(titles)

# ws.add_chart(chart, "E5")


################### Bubble Chart ########################
# from openpyxl.chart import Series, Reference, BubbleChart

# wb = Workbook()
# ws = wb.active

# rows = [
#     ("Number of Products", "Sales in USD", "Market share"),
#     (14, 12200, 15),
#     (20, 60000, 33),
#     (18, 24400, 10),
#     (22, 32000, 42),
#     (),
#     (12, 8200, 18),
#     (15, 50000, 30),
#     (19, 22400, 15),
#     (25, 25000, 50),
# ]

# for row in rows:
#     ws.append(row)

# chart = BubbleChart()
# chart.style = 18 # use a preset style

# # add the first series of data
# xvalues = Reference(ws, min_col=1, min_row=2, max_row=5)
# yvalues = Reference(ws, min_col=2, min_row=2, max_row=5)
# size = Reference(ws, min_col=3, min_row=2, max_row=5)
# series = Series(values=yvalues, xvalues=xvalues, zvalues=size, title="2013")
# chart.series.append(series)

# # add the second
# xvalues = Reference(ws, min_col=1, min_row=7, max_row=10)
# yvalues = Reference(ws, min_col=2, min_row=7, max_row=10)
# size = Reference(ws, min_col=3, min_row=7, max_row=10)
# series = Series(values=yvalues, xvalues=xvalues, zvalues=size, title="2014")
# chart.series.append(series)

# # place the chart starting in cell E1
# ws.add_chart(chart, "E1")


####################Line Charts#############################
# from datetime import date

# from openpyxl.chart import (
#     LineChart,
#     Reference,
# )
# from openpyxl.chart.axis import DateAxis

# wb = Workbook()
# ws = wb.active

# rows = [
#     ['Date', 'Batch 1', 'Batch 2', 'Batch 3'],
#     [date(2015,9, 1), 40, 30, 25],
#     [date(2015,9, 2), 40, 25, 30],
#     [date(2015,9, 3), 50, 30, 45],
#     [date(2015,9, 4), 30, 25, 40],
#     [date(2015,9, 5), 25, 35, 30],
#     [date(2015,9, 6), 20, 40, 35],
# ]

# for row in rows:
#     ws.append(row)

# c1 = LineChart()
# c1.title = "Line Chart"
# c1.style = 13
# c1.y_axis.title = 'Size'
# c1.x_axis.title = 'Test Number'

# data = Reference(ws, min_col=2, min_row=1, max_col=4, max_row=7)
# c1.add_data(data, titles_from_data=True)

# # Style the lines
# s1 = c1.series[0]
# s1.marker.symbol = "triangle"
# s1.marker.graphicalProperties.solidFill = "FF0000" # Marker filling
# s1.marker.graphicalProperties.line.solidFill = "FF0000" # Marker outline

# s1.graphicalProperties.line.noFill = True

# s2 = c1.series[1]
# s2.graphicalProperties.line.solidFill = "00AAAA"
# s2.graphicalProperties.line.dashStyle = "sysDot"
# s2.graphicalProperties.line.width = 100050 # width in EMUs

# s2 = c1.series[2]
# s2.smooth = True # Make the line smooth

# ws.add_chart(c1, "A10")

# from copy import deepcopy
# stacked = deepcopy(c1)
# stacked.grouping = "stacked"
# stacked.title = "Stacked Line Chart"
# ws.add_chart(stacked, "A27")

# percent_stacked = deepcopy(c1)
# percent_stacked.grouping = "percentStacked"
# percent_stacked.title = "Percent Stacked Line Chart"
# ws.add_chart(percent_stacked, "A44")

# # Chart with date axis
# c2 = LineChart()
# c2.title = "Date Axis"
# c2.style = 12
# c2.y_axis.title = "Size"
# c2.y_axis.crossAx = 500
# c2.x_axis = DateAxis(crossAx=100)
# c2.x_axis.number_format = 'd-mmm'
# c2.x_axis.majorTimeUnit = "days"
# c2.x_axis.title = "Date"

# c2.add_data(data, titles_from_data=True)
# dates = Reference(ws, min_col=1, min_row=2, max_row=7)
# c2.set_categories(dates)

# ws.add_chart(c2, "A61")


########################### 3D Line Chart #########################
# from datetime import date

# from openpyxl.chart import (
#     LineChart3D,
#     Reference,
# )
# from openpyxl.chart.axis import DateAxis

# wb = Workbook()
# ws = wb.active

# rows = [
#     ['Date', 'Batch 1', 'Batch 2', 'Batch 3'],
#     [date(2015,9, 1), 40, 30, 25],
#     [date(2015,9, 2), 40, 25, 30],
#     [date(2015,9, 3), 50, 30, 45],
#     [date(2015,9, 4), 30, 25, 40],
#     [date(2015,9, 5), 25, 35, 30],
#     [date(2015,9, 6), 20, 40, 35],
# ]

# for row in rows:
#     ws.append(row)

# c1 = LineChart3D()
# c1.title = "3D Line Chart"
# c1.legend = None
# c1.style = 15
# c1.y_axis.title = 'Size'
# c1.x_axis.title = 'Test Number'

# data = Reference(ws, min_col=2, min_row=1, max_col=4, max_row=7)
# c1.add_data(data, titles_from_data=True)

# ws.add_chart(c1, "A10")

# para guardar en el archivo
wb.save(file_name)